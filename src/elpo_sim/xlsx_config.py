def update_dict_from_xlsx(cfg, xlsx_path, sheet_name=None):
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        key_path = str(row[0]).strip()
        value = row[1]
        if value is None:
            continue
        node = cfg
        parts = key_path.split(".")
        for part in parts[:-1]:
            node = node.setdefault(part, {})
        node[parts[-1]] = value
    return cfg

